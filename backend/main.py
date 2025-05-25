from database import *
from sqlalchemy.orm import Session

from fastapi import Depends, FastAPI, Body
from fastapi.responses import JSONResponse, FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font

import os


Base.metadata.create_all(bind=engine)

app = FastAPI()

# Для запуска локально раскомментировать эту строку и выполнить npm run build
# app.mount("/assets", StaticFiles(directory='../frontend/dist/assets'), name='static')

# Для запуска на удаленном сервере раскомментировать эту строку
app.mount("/assets", StaticFiles(directory='/usr/share/nginx/html/assets'), name='static')


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


@app.get("/")
def main():
    # При локальном запуске раскомментировать эту строку
    # return FileResponse("../frontend/dist/index.html")

    # При запуске на удаленном сервере раскомментировать эту строку
    return FileResponse("/usr/share/nginx/html/index.html")


@app.get("/api/trainerList")
def get_trainers_workouts(db: Session = Depends(get_db)):
    return db.query(Trainer).all()


@app.get("/api/workoutList")
def get_workouts(db: Session = Depends(get_db)):
    return db.query(Workout).all()


@app.get("/api/get-table/{month_year}")
def get_table(month_year, db: Session = Depends(get_db)):
    if db.query(WorkoutTable).filter(WorkoutTable.month_year == month_year).first() != None:
        return db.query(WorkoutTable).filter(WorkoutTable.month_year == month_year).first()
    else:
        return JSONResponse({"message": "Расписания на такую дату не существует"})


@app.get("/api/download-schedule")
def download_file():
    file_path = os.path.abspath('../frontend/src/Расписание_тренировок.xlsx')

    if os.path.exists(file_path):
        return FileResponse(
            path=file_path,
            filename='Расписание_тренировок.xlsx',
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    else:
        return {"error": "Файл не найден"}


@app.post("/api/trainerList")
def add_trainer(data=Body(), db: Session = Depends(get_db)):
    if data["name"] != "":
        trainer = Trainer(name=data["name"])
        db.add(trainer)
        db.commit()
        db.refresh(trainer)
    return trainer


@app.post("/api/workoutList")
def add_workout(data=Body(), db: Session = Depends(get_db)):
    if data["name"] != "":
        workout = Workout(name=data["name"])
        db.add(workout)
        db.commit()
        db.refresh(workout)
    return workout


@app.post("/api/save-table/{month_year}")
def add_workout_table_data(month_year: str, data=Body(), db: Session = Depends(get_db)):
    if db.query(WorkoutTable).filter(WorkoutTable.month_year == month_year).first() is None:
        table_data = WorkoutTable(month_year=month_year, table_data=data["rows"])
        db.add(table_data)
        db.commit()
        db.refresh(table_data)

        excel_writing(month_year, data["rows"])

        return table_data
    else:
        return JSONResponse(content={"message": "Расписание на такой месяц уже сущетствует"})


@app.put("/api/save-table/{month_year}")
def edit_workout_table_data(month_year: str, data=Body(), db: Session = Depends(get_db)):
    table_data = db.query(WorkoutTable).filter(WorkoutTable.month_year == month_year).first()

    table_data.table_data = data["rows"]
    db.commit()
    db.refresh(table_data)

    excel_writing(month_year, data["rows"])

    return table_data


@app.delete("/api/trainerList/{id}")
def delete_trainer(id, db: Session = Depends(get_db)):
    trainer = db.query(Trainer).filter(Trainer.id == id).first()

    if trainer is None:
        return JSONResponse(status_code=404, content={"message": "Trainer not found"})

    db.delete(trainer)
    db.commit()
    return trainer


@app.delete("/api/workoutList/{id}")
def delete_workout(id, db: Session = Depends(get_db)):
    workout = db.query(Workout).filter(Workout.id == id).first()

    if workout is None:
        return JSONResponse(status_code=404, content={"message": "Trainer not found"})

    db.delete(workout)
    db.commit()
    return workout


def excel_writing(month_year, workbook_data):
    file_path = os.path.abspath('../frontend/src/Расписание_тренировок.xlsx')
    wb = load_workbook(file_path)

    if month_year not in wb.sheetnames:
        wb.create_sheet(month_year)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    workbook_list = wb[month_year]

    workbook_list.row_dimensions[1].height = 20

    headers = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье', '']

    row_index = 0
    cell_index = 0

    for data in workbook_data:
        for key, value in data.items():
            if key == "cells":
                for cell in value:

                    if cell_index < 7:
                        cell_index += 1
                    else:
                        cell_index = 1

                    headers_coordinate = workbook_list['B1':'H1'][0][cell_index - 1].coordinate
    
                    workbook_list[headers_coordinate] = headers[cell_index - 1]
                    workbook_list[headers_coordinate].border = thin_border
                    workbook_list[headers_coordinate].font = Font(bold=True)

                    workbook_coordinate = workbook_list['A2':'H13'][row_index][cell_index].coordinate

                    workbook_list[workbook_coordinate].border = thin_border
                    workbook_list[workbook_coordinate] = cell['text']
                    workbook_list[workbook_coordinate].fill = PatternFill('solid', fgColor=cell['color'][1:])

                    workbook_list.column_dimensions[workbook_coordinate[0]].width = 25
            else:
                time_header_coordinate = workbook_list['A1':'A1'][0][0].coordinate

                workbook_list[time_header_coordinate] = headers[7]
                workbook_list[time_header_coordinate].border = thin_border
                workbook_list.column_dimensions[time_header_coordinate[0]].width = 15

                time_cell_coord = workbook_list['A2':'H13'][row_index][0].coordinate

                workbook_list[time_cell_coord] = data[key]
                workbook_list[time_cell_coord].border = thin_border
                workbook_list[time_cell_coord].font = Font(bold=True)

        workbook_list.row_dimensions[row_index + 2].height = 40

        row_index += 1

    wb.save(file_path)
    wb.close()
